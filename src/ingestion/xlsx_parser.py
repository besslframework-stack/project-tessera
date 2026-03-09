"""Parse XLSX files into LlamaIndex Documents."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from llama_index.core.schema import Document

logger = logging.getLogger(__name__)


def parse_xlsx_file(file_path: Path) -> list["Document"]:
    """Parse an XLSX file into Documents. Each sheet, each row becomes a Document."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed, skipping XLSX: %s", file_path)
        return []

    from llama_index.core.schema import Document

    documents = []
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    except Exception as exc:
        logger.error("Failed to open XLSX %s: %s", file_path, exc)
        return []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_list = list(ws.iter_rows(values_only=True))
        if not rows_list or len(rows_list) < 2:
            continue

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows_list[0])]

        for row_idx, row in enumerate(rows_list[1:], start=1):
            text_parts = []
            for h, v in zip(headers, row):
                if v is not None and str(v).strip():
                    text_parts.append(f"{h}: {str(v).strip()}")

            text = " | ".join(text_parts)
            if not text.strip():
                continue

            documents.append(Document(
                text=text,
                metadata={
                    "source_path": str(file_path),
                    "file_name": file_path.name,
                    "file_type": "xlsx",
                    "sheet_name": sheet_name,
                    "row_index": row_idx,
                },
            ))

    wb.close()
    return documents


def format_xlsx_as_table(file_path: Path) -> str:
    """Read an XLSX file and return its contents as markdown tables (one per sheet)."""
    try:
        import openpyxl
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl"

    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    except Exception as exc:
        return f"Failed to open {file_path.name}: {exc}"

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_list = list(ws.iter_rows(values_only=True))
        if not rows_list:
            continue

        headers = [str(h) if h is not None else "" for h in rows_list[0]]
        lines = [f"### {sheet_name}", ""]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

        for row in rows_list[1:]:
            vals = [
                str(v).replace("|", "\\|").replace("\n", " ") if v is not None else ""
                for v in row
            ]
            lines.append("| " + " | ".join(vals) + " |")

        parts.append("\n".join(lines))

    wb.close()
    return "\n\n".join(parts) if parts else f"No data found in {file_path.name}"
